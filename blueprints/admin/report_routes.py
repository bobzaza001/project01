import io
from datetime import datetime
from flask import render_template, request, send_file, flash, redirect, url_for
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from models import db, BorrowRequest, RepairRequest, Equipment, User, get_local_now
from utils import admin_required
from . import admin_bp

THAI_MONTHS = [
    "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
    "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
]

def format_thai_date(dt, include_time=True):
    if not dt:
        return "-"
    thai_year = dt.year + 543
    month_name = THAI_MONTHS[dt.month]
    if include_time:
        return f"{dt.day} {month_name} {thai_year} {dt.strftime('%H:%M น.')}"
    return f"{dt.day} {month_name} {thai_year}"

def style_excel_sheet(ws, title_text, subtitle_text, headers):
    # Colors
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Cordia New", size=14, bold=True, color="FFFFFF")
    title_font = Font(name="Cordia New", size=18, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Cordia New", size=13, italic=True, color="475569")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title_text).font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=subtitle_text).font = subtitle_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    # Empty Row 3
    
    # Header Row 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 25
    return thin_border

def autofit_columns(ws, max_cols):
    for col in range(1, max_cols + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(4, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


# ==========================================
# 1. EXCEL: ประวัติการยืม-คืน (Borrow History)
# ==========================================
@admin_bp.route('/export/borrow-history')
@admin_required
def export_borrow_history_excel():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    status_filter = request.args.get('status', 'all')
    
    query = BorrowRequest.query
    
    if month and year:
        query = query.filter(
            db.extract('month', BorrowRequest.requested_at) == month,
            db.extract('year', BorrowRequest.requested_at) == year
        )
        period_text = f"ประจำเดือน {THAI_MONTHS[month]} พ.ศ. {year + 543}"
    else:
        period_text = "ประวัติทั้งหมดตั้งแต่เริ่มระบบ"
        
    if status_filter and status_filter != 'all':
        query = query.filter_by(status=status_filter)
        
    records = query.order_by(BorrowRequest.requested_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ประวัติการยืม-คืน"
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "ลำดับ", "เลขที่คำขอ", "วันที่ยื่นคำขอ", "ผู้ยื่นคำขอ", "อีเมล / รหัสนักศึกษา",
        "ชื่ออุปกรณ์ / ครุภัณฑ์", "รหัสครุภัณฑ์", "ประเภท", "จำนวน/วัน",
        "กำหนดคืน", "วันที่คืนจริง", "สถานะ", "สภาพส่งคืน", "หมายเหตุ / เหตุผล"
    ]
    
    title = "รายงานสรุปประวัติการยืม-คืนครุภัณฑ์และเบิกวัสดุ — วิทยาลัยเทคโนโลยีพณิชยการอยุธยา"
    subtitle = f"{period_text} | ข้อมูล ณ วันที่ {format_thai_date(get_local_now())}"
    thin_border = style_excel_sheet(ws, title, subtitle, headers)
    
    status_map = {
        'pending': 'รออนุมัติ',
        'approved': 'กำลังยืม',
        'return_pending': 'รอแอดมินรับคืน',
        'returned': 'คืนแล้ว',
        'rejected': 'ปฏิเสธ'
    }
    
    data_font = Font(name="Cordia New", size=13)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    current_row = 5
    for idx, r in enumerate(records, 1):
        req_type = "วัสดุสิ้นเปลือง" if r.equipment and r.equipment.item_type == 'consumable' else "ครุภัณฑ์"
        qty_days = f"{r.quantity} ชิ้น" if r.equipment and r.equipment.item_type == 'consumable' else f"{r.borrow_days or 3} วัน"
        
        status_th = status_map.get(r.status, r.status)
        damage_th = "ชำรุด" if r.damage_status == 'damaged' else ("ปกติ" if r.status == 'returned' else "-")
        note = r.damage_note if r.damage_status == 'damaged' else (r.warning_message or "-")
        
        row_values = [
            idx,
            r.id,
            r.requested_at.strftime('%d/%m/%Y %H:%M') if r.requested_at else "-",
            r.requester.full_name if r.requester else "-",
            r.requester.email or r.requester.username if r.requester else "-",
            r.equipment.name if r.equipment else "-",
            r.equipment.equipment_code if r.equipment else "-",
            req_type,
            qty_days,
            r.return_due_datetime.strftime('%d/%m/%Y') if r.return_due_datetime else "-",
            r.returned_at.strftime('%d/%m/%Y %H:%M') if r.returned_at else "-",
            status_th,
            damage_th,
            note
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if current_row % 2 == 0:
                cell.fill = zebra_fill
            if col_idx in [1, 2, 8, 9, 10, 11, 12, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    autofit_columns(ws, len(headers))
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"รายงานประวัติการยืมคืน_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ==========================================
# 2. EXCEL: รายการแจ้งซ่อม/ชำรุด (Repairs & Damaged)
# ==========================================
@admin_bp.route('/export/repairs')
@admin_required
def export_repairs_excel():
    status_filter = request.args.get('status', 'all')
    
    query = RepairRequest.query
    if status_filter and status_filter != 'all':
        query = query.filter_by(status=status_filter)
        
    records = query.order_by(RepairRequest.reported_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานแจ้งซ่อมและชำรุด"
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "ลำดับ", "เลขที่แจ้งซ่อม", "วันที่แจ้ง", "ชื่อครุภัณฑ์", "รหัสครุภัณฑ์",
        "สถานที่จัดเก็บ", "ผู้แจ้งเรื่อง", "รายละเอียดอาการชำรุด", "สถานะการซ่อม",
        "วันที่ซ่อมเสร็จ", "บันทึกผลการซ่อมแซม"
    ]
    
    title = "รายงานสรุปรายการครุภัณฑ์ชำรุดและการแจ้งซ่อมบำรุง — วิทยาลัยเทคโนโลยีพณิชยการอยุธยา"
    subtitle = f"ข้อมูล ณ วันที่ {format_thai_date(get_local_now())} | รวมทั้งหมด {len(records)} รายการ"
    thin_border = style_excel_sheet(ws, title, subtitle, headers)
    
    status_map = {
        'pending': 'รอดำเนินการ',
        'in_progress': 'กำลังดำเนินการซ่อม',
        'completed': 'ซ่อมเสร็จสิ้น'
    }
    
    data_font = Font(name="Cordia New", size=13)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    current_row = 5
    for idx, r in enumerate(records, 1):
        loc = "-"
        if r.equipment and r.equipment.room_ref:
            room = r.equipment.room_ref
            floor_name = room.floor.name if room.floor else ""
            bldg_name = room.floor.building.name if room.floor and room.floor.building else ""
            loc = f"{bldg_name} {floor_name} ({room.name})".strip()
            
        row_values = [
            idx,
            r.id,
            r.reported_at.strftime('%d/%m/%Y %H:%M') if r.reported_at else "-",
            r.equipment.name if r.equipment else "-",
            r.equipment.equipment_code if r.equipment else "-",
            loc,
            r.reporter.full_name if r.reporter else "-",
            r.issue_description,
            status_map.get(r.status, r.status),
            r.resolved_at.strftime('%d/%m/%Y %H:%M') if r.resolved_at else "-",
            r.admin_note or "-"
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if current_row % 2 == 0:
                cell.fill = zebra_fill
            if col_idx in [1, 2, 3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    autofit_columns(ws, len(headers))
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"รายงานการแจ้งซ่อมครุภัณฑ์_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ==========================================
# 3. EXCEL: คลังครุภัณฑ์และวัสดุ (Inventory Stock)
# ==========================================
@admin_bp.route('/export/inventory')
@admin_required
def export_inventory_excel():
    items = Equipment.query.order_by(Equipment.id.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ยอดคลังครุภัณฑ์และวัสดุ"
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "ลำดับ", "รหัสครุภัณฑ์/วัสดุ", "ชื่อรายการ", "หมวดหมู่", "ประเภท",
        "สถานที่จัดเก็บ (ตึก/ชั้น/ห้อง)", "จำนวนทั้งหมด", "จำนวนพร้อมใช้งาน",
        "จำนวนที่ถูกยืม/ใช้งาน", "สถานะ", "สิทธิ์ยืมกลับบ้าน"
    ]
    
    title = "รายงานสรุปยอดคงคลังครุภัณฑ์และวัสดุสิ้นเปลือง — วิทยาลัยเทคโนโลยีพณิชยการอยุธยา"
    subtitle = f"ข้อมูล ณ วันที่ {format_thai_date(get_local_now())} | รวม {len(items)} รายการ"
    thin_border = style_excel_sheet(ws, title, subtitle, headers)
    
    status_map = {
        'available': 'พร้อมใช้งาน',
        'borrowed': 'ถูกยืมทั้งหมด',
        'maintenance': 'ส่งซ่อมบำรุง',
        'disposed': 'ตัดจำหน่ายแล้ว'
    }
    
    data_font = Font(name="Cordia New", size=13)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    current_row = 5
    for idx, eq in enumerate(items, 1):
        loc = "-"
        if eq.room_ref:
            room = eq.room_ref
            floor_name = room.floor.name if room.floor else ""
            bldg_name = room.floor.building.name if room.floor and room.floor.building else ""
            loc = f"{bldg_name} {floor_name} ({room.name})".strip()
            
        in_use = eq.total_quantity - eq.available_quantity
        
        row_values = [
            idx,
            eq.equipment_code,
            eq.name,
            eq.category or "ทั่วไป",
            "วัสดุสิ้นเปลือง" if eq.item_type == 'consumable' else "ครุภัณฑ์",
            loc,
            eq.total_quantity,
            eq.available_quantity,
            in_use if in_use >= 0 else 0,
            status_map.get(eq.status, eq.status),
            "อนุญาต" if eq.is_borrowable else "ไม่อนุญาต"
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if current_row % 2 == 0:
                cell.fill = zebra_fill
            if col_idx in [1, 2, 4, 5, 7, 8, 9, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    autofit_columns(ws, len(headers))
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"รายงานยอดคลังพัสดุครุภัณฑ์_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ==========================================
# 4. PRINT / PDF: รายงานการยืม-คืน (Borrow Print)
# ==========================================
@admin_bp.route('/reports/borrow-print')
@admin_required
def report_borrow_print():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    status_filter = request.args.get('status', 'all')
    
    query = BorrowRequest.query
    if month and year:
        query = query.filter(
            db.extract('month', BorrowRequest.requested_at) == month,
            db.extract('year', BorrowRequest.requested_at) == year
        )
        period_text = f"ประจำเดือน {THAI_MONTHS[month]} พ.ศ. {year + 543}"
    else:
        period_text = "ประวัติรายการทั้งหมด"
        
    if status_filter and status_filter != 'all':
        query = query.filter_by(status=status_filter)
        
    records = query.order_by(BorrowRequest.requested_at.desc()).all()
    
    stats = {
        'total': len(records),
        'returned': sum(1 for r in records if r.status == 'returned'),
        'active': sum(1 for r in records if r.status in ['approved', 'return_pending']),
        'rejected': sum(1 for r in records if r.status == 'rejected'),
        'damaged': sum(1 for r in records if r.damage_status == 'damaged')
    }
    
    return render_template('report_print.html',
                           report_type='borrow',
                           report_title='รายงานสรุปประวัติการยืม-คืนครุภัณฑ์และเบิกจ่ายวัสดุ',
                           period_text=period_text,
                           records=records,
                           stats=stats,
                           generated_date=format_thai_date(get_local_now()))


# ==========================================
# 5. PRINT / PDF: รายงานแจ้งซ่อม/ชำรุด (Repairs Print)
# ==========================================
@admin_bp.route('/reports/repairs-print')
@admin_required
def report_repairs_print():
    records = RepairRequest.query.order_by(RepairRequest.reported_at.desc()).all()
    
    stats = {
        'total': len(records),
        'completed': sum(1 for r in records if r.status == 'completed'),
        'in_progress': sum(1 for r in records if r.status == 'in_progress'),
        'pending': sum(1 for r in records if r.status == 'pending')
    }
    
    return render_template('report_print.html',
                           report_type='repairs',
                           report_title='รายงานสรุปรายการครุภัณฑ์ชำรุดและการแจ้งซ่อมบำรุง',
                           period_text='รายการแจ้งซ่อมทั้งหมดในระบบ',
                           records=records,
                           stats=stats,
                           generated_date=format_thai_date(get_local_now()))


# ==========================================
# 6. PRINT / PDF: รายงานยอดคลังพัสดุ (Inventory Print)
# ==========================================
@admin_bp.route('/reports/inventory-print')
@admin_required
def report_inventory_print():
    items = Equipment.query.order_by(Equipment.id.asc()).all()
    
    stats = {
        'total_items': len(items),
        'total_qty': sum(eq.total_quantity for eq in items),
        'available_qty': sum(eq.available_quantity for eq in items),
        'borrowed_qty': sum(eq.total_quantity - eq.available_quantity for eq in items if eq.total_quantity > eq.available_quantity),
        'maintenance_count': sum(1 for eq in items if eq.status == 'maintenance')
    }
    
    return render_template('report_print.html',
                           report_type='inventory',
                           report_title='รายงานสรุปยอดคงคลังครุภัณฑ์และวัสดุประจำห้องปฏิบัติการ',
                           period_text='ข้อมูลสถานะคลังพัสดุล่าสุด',
                           items=items,
                           stats=stats,
                           generated_date=format_thai_date(get_local_now()))
