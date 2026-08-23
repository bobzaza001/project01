import io
from datetime import datetime, timedelta
from flask import render_template, request, send_file, flash, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import db, AuditLog, User, get_local_now
from utils import admin_required
from . import admin_bp

@admin_bp.route('/audit-logs')
@admin_required
def audit_logs():
    """หน้าตรวจสอบประวัติการแก้ไขและจัดการข้อมูลระบบ (Audit Logs)"""
    category = request.args.get('category', 'all')
    search_q = request.args.get('q', '').strip()
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 30

    query = AuditLog.query

    # กรองตามหมวดหมู่
    if category in ('equipment', 'location'):
        query = query.filter(AuditLog.category == category)

    # กรองตามคำค้นหา
    if search_q:
        search_pattern = f"%{search_q}%"
        query = query.outerjoin(User, AuditLog.user_id == User.id).filter(
            db.or_(
                AuditLog.target_name.ilike(search_pattern),
                AuditLog.details.ilike(search_pattern),
                AuditLog.action.ilike(search_pattern),
                User.full_name.ilike(search_pattern),
                User.username.ilike(search_pattern)
            )
        )

    # กรองตามวันที่
    if date_from_str:
        try:
            df = datetime.strptime(date_from_str, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at >= df)
        except ValueError:
            pass

    if date_to_str:
        try:
            dt = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.created_at < dt)
        except ValueError:
            pass

    # เรียงลำดับจากล่าสุดไปเก่าสุด
    query = query.order_by(AuditLog.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # สถิติภาพรวม
    today_start = get_local_now().replace(hour=0, minute=0, second=0, microsecond=0)
    total_logs = AuditLog.query.count()
    total_equipment_logs = AuditLog.query.filter_by(category='equipment').count()
    total_location_logs = AuditLog.query.filter_by(category='location').count()
    today_logs = AuditLog.query.filter(AuditLog.created_at >= today_start).count()

    return render_template(
        'admin_audit_logs.html',
        logs=pagination.items,
        pagination=pagination,
        category=category,
        search_q=search_q,
        date_from=date_from_str,
        date_to=date_to_str,
        total_logs=total_logs,
        total_equipment_logs=total_equipment_logs,
        total_location_logs=total_location_logs,
        today_logs=today_logs
    )


@admin_bp.route('/export/audit-logs')
@admin_required
def export_audit_logs():
    """ส่งออกประวัติการแก้ไขและจัดการข้อมูลเป็นไฟล์ Excel (.xlsx)"""
    category = request.args.get('category', 'all')
    search_q = request.args.get('q', '').strip()
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()

    query = AuditLog.query

    if category in ('equipment', 'location'):
        query = query.filter(AuditLog.category == category)

    if search_q:
        search_pattern = f"%{search_q}%"
        query = query.outerjoin(User, AuditLog.user_id == User.id).filter(
            db.or_(
                AuditLog.target_name.ilike(search_pattern),
                AuditLog.details.ilike(search_pattern),
                AuditLog.action.ilike(search_pattern),
                User.full_name.ilike(search_pattern)
            )
        )

    if date_from_str:
        try:
            df = datetime.strptime(date_from_str, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at >= df)
        except ValueError:
            pass

    if date_to_str:
        try:
            dt = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.created_at < dt)
        except ValueError:
            pass

    logs = query.order_by(AuditLog.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit_Logs"
    ws.views.sheetView[0].showGridLines = True

    # Styling
    title_font = Font(name="Sarabun", size=15, bold=True, color="0F172A")
    subtitle_font = Font(name="Sarabun", size=10, italic=True, color="475569")
    header_font = Font(name="Sarabun", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Sarabun", size=10, color="1E293B")
    bold_data_font = Font(name="Sarabun", size=10, bold=True, color="0F172A")

    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Banner
    ws.merge_cells('A1:G1')
    ws['A1'] = "รายงานบันทึกประวัติการแก้ไขและจัดการข้อมูลระบบ (System Audit Log)"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    gen_time_str = get_local_now().strftime('%d/%m/%Y %H:%M น.')
    ws.merge_cells('A2:G2')
    ws['A2'] = f"วิทยาลัยเทคโนโลยีพณิชยการอยุธยา (ATCC) | ข้อมูล ณ วันที่ {gen_time_str} | ทั้งหมด {len(logs)} รายการ"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        "ลำดับ",
        "วัน-เวลาที่ดำเนินการ",
        "ผู้ดำเนินการ (Admin)",
        "หมวดหมู่",
        "กิจกรรม (Action)",
        "รายการเป้าหมาย",
        "รายละเอียดการเปลี่ยนแปลง"
    ]
    
    ws.append([]) # Blank row 3
    ws.row_dimensions[3].height = 10
    
    ws.append(headers)
    ws.row_dimensions[4].height = 28
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    action_labels = {
        'ADD_EQUIPMENT': 'เพิ่มครุภัณฑ์ใหม่',
        'EDIT_EQUIPMENT': 'แก้ไขข้อมูลครุภัณฑ์',
        'DELETE_EQUIPMENT': 'ลบครุภัณฑ์',
        'DISPOSE_EQUIPMENT': 'ตัดจำหน่ายครุภัณฑ์',
        'ADD_BUILDING': 'เพิ่มอาคารใหม่',
        'EDIT_BUILDING': 'แก้ไขชื่ออาคาร',
        'DELETE_BUILDING': 'ลบอาคาร',
        'ADD_FLOOR': 'เพิ่มชั้นใหม่',
        'EDIT_FLOOR': 'แก้ไขชื่อชั้น',
        'DELETE_FLOOR': 'ลบชั้น',
        'ADD_ROOM': 'เพิ่มห้องใหม่',
        'ADD_ROOM_BULK': 'เพิ่มห้องแบบกลุ่ม',
        'EDIT_ROOM': 'แก้ไขชื่อห้อง',
        'DELETE_ROOM': 'ลบห้อง'
    }

    row_num = 5
    for idx, log in enumerate(logs, start=1):
        dt_str = log.created_at.strftime('%d/%m/%Y %H:%M น.') if log.created_at else '-'
        admin_name = log.user.full_name if log.user else 'ระบบอัตโนมัติ'
        cat_str = 'ครุภัณฑ์/วัสดุ' if log.category == 'equipment' else 'อาคาร/ชั้น/ห้อง'
        action_str = action_labels.get(log.action, log.action)

        row_data = [
            idx,
            dt_str,
            admin_name,
            cat_str,
            action_str,
            log.target_name,
            log.details or '-'
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 24
        current_fill = alt_fill if idx % 2 == 0 else white_fill

        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = bold_data_font if col_num in (1, 3, 5) else data_font
            cell.fill = current_fill
            cell.border = thin_border
            
            if col_num in (1, 2, 4, 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        row_num += 1

    # Adjust Column Widths
    col_widths = {
        'A': 8,
        'B': 22,
        'C': 24,
        'D': 18,
        'E': 22,
        'F': 35,
        'G': 55
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"System_Audit_Logs_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
